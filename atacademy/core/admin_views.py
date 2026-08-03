import io
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook

from .models import Enquiry, CallbackRequest, BrochureRequest, RecruiterContact


def _make_excel(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_response(buf, filename):
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_member_required
def export_dashboard(request):
    context = {
        'enquiry_count': Enquiry.objects.count(),
        'callback_count': CallbackRequest.objects.count(),
        'brochure_count': BrochureRequest.objects.count(),
        'recruiter_count': RecruiterContact.objects.count(),
        'total': Enquiry.objects.count() + CallbackRequest.objects.count() + BrochureRequest.objects.count() + RecruiterContact.objects.count(),
    }
    return render(request, 'admin/export_dashboard.html', context)


@staff_member_required
def export_enquiries(request):
    qs = Enquiry.objects.all().order_by('-created_at')
    headers = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Qualification', 'Date', 'Read']
    rows = [[e.name, e.email, e.phone, e.course, e.branch, e.qualification,
             e.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if e.is_read else 'No'] for e in qs]
    buf = _make_excel('Enquiries', headers, rows)
    return _make_response(buf, f'enquiries_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')


@staff_member_required
def export_callbacks(request):
    qs = CallbackRequest.objects.all().order_by('-created_at')
    headers = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Date', 'Read']
    rows = [[c.name, c.email, c.phone, c.course, c.branch,
             c.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if c.is_read else 'No'] for c in qs]
    buf = _make_excel('Callback Requests', headers, rows)
    return _make_response(buf, f'callbacks_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')


@staff_member_required
def export_brochures(request):
    qs = BrochureRequest.objects.select_related('course').all().order_by('-created_at')
    headers = ['Name', 'Email', 'Phone', 'Course', 'Date', 'Read']
    rows = [[b.name, b.email, b.phone, b.course.name if b.course else '',
             b.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if b.is_read else 'No'] for b in qs]
    buf = _make_excel('Brochure Requests', headers, rows)
    return _make_response(buf, f'brochures_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')


@staff_member_required
def export_recruiters(request):
    qs = RecruiterContact.objects.all().order_by('-created_at')
    headers = ['Name', 'Email', 'Phone', 'Company', 'Designation', 'Date', 'Read']
    rows = [[r.name, r.email, r.phone, r.company_name, r.designation,
             r.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if r.is_read else 'No'] for r in qs]
    buf = _make_excel('Recruiter Contacts', headers, rows)
    return _make_response(buf, f'recruiters_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')


@staff_member_required
def export_all(request):
    wb = Workbook()

    ws = wb.active
    ws.title = 'Enquiries'
    headers = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Qualification', 'Date', 'Read']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)
    for row_idx, e in enumerate(Enquiry.objects.all().order_by('-created_at'), 2):
        for col_idx, val in enumerate([e.name, e.email, e.phone, e.course, e.branch, e.qualification,
                                       e.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if e.is_read else 'No'], 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val else '')

    ws2 = wb.create_sheet('Callback Requests')
    headers2 = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Date', 'Read']
    for col_idx, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_idx, value=h).font = ws2.cell(row=1, column=col_idx).font.copy(bold=True)
    for row_idx, c in enumerate(CallbackRequest.objects.all().order_by('-created_at'), 2):
        for col_idx, val in enumerate([c.name, c.email, c.phone, c.course, c.branch,
                                       c.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if c.is_read else 'No'], 1):
            ws2.cell(row=row_idx, column=col_idx, value=str(val) if val else '')

    ws3 = wb.create_sheet('Brochure Requests')
    headers3 = ['Name', 'Email', 'Phone', 'Course', 'Date', 'Read']
    for col_idx, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col_idx, value=h).font = ws3.cell(row=1, column=col_idx).font.copy(bold=True)
    for row_idx, b in enumerate(BrochureRequest.objects.select_related('course').all().order_by('-created_at'), 2):
        for col_idx, val in enumerate([b.name, b.email, b.phone, b.course.name if b.course else '',
                                       b.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if b.is_read else 'No'], 1):
            ws3.cell(row=row_idx, column=col_idx, value=str(val) if val else '')

    ws4 = wb.create_sheet('Recruiter Contacts')
    headers4 = ['Name', 'Email', 'Phone', 'Company', 'Designation', 'Date', 'Read']
    for col_idx, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col_idx, value=h).font = ws4.cell(row=1, column=col_idx).font.copy(bold=True)
    for row_idx, r in enumerate(RecruiterContact.objects.all().order_by('-created_at'), 2):
        for col_idx, val in enumerate([r.name, r.email, r.phone, r.company_name, r.designation,
                                       r.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if r.is_read else 'No'], 1):
            ws4.cell(row=row_idx, column=col_idx, value=str(val) if val else '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _make_response(buf, f'all_data_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')
